"""Extract public assessment definitions from MHSDS ETOS v4.1/v5/v6 workbooks.

Run with the bundled openpyxl runtime. Inputs are VERSION=PATH pairs, oldest
first. This reads specification metadata only, never warehouse clinical data.
"""

import argparse
import csv
import re
from pathlib import Path

import openpyxl


FIELDS = [
    'concept_code', 'assessment_tool_name', 'assessment_description',
    'published_value', 'response_description', 'decimal_places',
    'collection_start_date', 'specification_version', 'source_row',
]


def text(value):
    return '' if value is None else ' '.join(str(value).split())


def extract(version, path):
    if not version.startswith(('4.1', '5.', '6.')):
        raise ValueError('Only ETOS v4.1, v5 and v6 worksheet layouts are supported')
    book = openpyxl.load_workbook(path, data_only=True)
    sheet = book['MH Assessment Scales']
    # Older ETOS sheets merge tool, concept and precision cells across responses.
    merged = {}
    for area in sheet.merged_cells.ranges:
        value = sheet.cell(area.min_row, area.min_col).value
        for row in range(area.min_row, area.max_row + 1):
            for col in range(area.min_col, area.max_col + 1):
                merged[row, col] = value

    def cell(row, col):
        return merged.get((row, col), sheet.cell(row, col).value)

    tool, description, active, inactive, value, label, decimals, start = (
        (2, 3, 4, 11, 5, 6, 7, 10) if version.startswith('6')
        else (1, 2, 3, 4, 5, 6, 12, 13)
    )
    for row in range(1, sheet.max_row + 1):
        active_code = text(cell(row, active))
        if not re.fullmatch(r'[0-9]{6,18}', active_code):
            continue
        codes = {active_code} | set(re.findall(r'\b[0-9]{6,18}\b', text(cell(row, inactive))))
        published_value = text(cell(row, value))
        if not published_value:
            continue
        start_date = cell(row, start)
        for code in sorted(codes):
            yield dict(zip(FIELDS, [
                code, text(cell(row, tool)), text(cell(row, description)),
                published_value, text(cell(row, label)), text(cell(row, decimals)),
                start_date.date().isoformat() if hasattr(start_date, 'date') else '',
                version, row,
            ]))
    book.close()


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--spec', action='append', required=True, metavar='VERSION=PATH')
    parser.add_argument('--output', type=Path, required=True)
    args = parser.parse_args()
    definitions = {}
    for spec in args.spec:
        version, path = spec.split('=', 1)
        for row in extract(version, path):
            key = row['concept_code'], row['published_value'], version
            previous = definitions.get(key)
            if previous and any(previous[f] != row[f] for f in FIELDS if f != 'source_row'):
                raise ValueError(f'Conflicting public definitions for {key}')
            definitions[key] = row
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open('w', encoding='utf-8', newline='') as stream:
        writer = csv.DictWriter(stream, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(definitions[key] for key in sorted(definitions))
    print(f'Extracted {len(definitions)} public definitions for '
          f'{len({key[0] for key in definitions})} current or historical concepts.')


if __name__ == '__main__':
    main()
